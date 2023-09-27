from tkinter import *
from tkinter.ttk import *
import tkinter as tk
import tkinter.ttk as ttk
from PIL import ImageTk, Image
import webbrowser
from openpyxl import load_workbook

#carrega a planilha
caminho = './assets/spreadsheets/LEIAUTE PADRAO.xlsx'
caminho_icon = "./assets/images/icon-removebg-preview.ico"
caminho_logo = "./assets/images/LOGO2.png"
caminho_tema = './assets/themes/azure/azure.tcl'

arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active

def linkContratoSocial():
    contratoSocial = planilha1.cell(row = 8, column = 3).value
    webbrowser.open(contratoSocial)
def linkContratosDiversos():
    contratoDiverso = planilha1.cell(row = 9, column = 3).value
    webbrowser.open(contratoDiverso)
def linkCertidoes():
    certidao = planilha1.cell(row = 10, column = 3).value
    webbrowser.open(certidao)
def linkModelos():
    modelos = planilha1.cell(row = 11, column = 3).value
    webbrowser.open(modelos)

def jan_gestaoEmpresa():
    #cria a janela
    gestaoEmpresa = tk.Toplevel()
    gestaoEmpresa.grab_set()
    gestaoEmpresa.title("Gestão da Empresa")
    gestaoEmpresa.geometry("400x360")
    gestaoEmpresa.iconbitmap(caminho_icon)
    #gestaoEmpresa.resizable(False, False)
    
    #carrega a logo
    img = ImageTk.PhotoImage(Image.open(caminho_logo))
    panel2 = tk.Label(gestaoEmpresa, image = img)
    panel2.grid(row = 0, columnspan = 3, column = 0, sticky = N, padx = 55, pady = 30)
    panel2.image = img
    
    #criando botões
    contratoSocial = ttk.Button(gestaoEmpresa, text = 'Contrato Social', width = 23, style = "Accentbutton", command = linkContratoSocial)
    contratoSocial.grid(row = 1, column = 0, sticky = W, padx = 27, pady = 10)
    contratosDiversos = ttk.Button(gestaoEmpresa, text = 'Contratos Diversos', width = 23, style = "Accentbutton", command = linkContratosDiversos)
    contratosDiversos.grid(row = 1, column = 2, sticky = W, pady = 10)
    certidoes = ttk.Button(gestaoEmpresa, text = 'Certidões', width = 23, style = "Accentbutton", command = linkCertidoes)
    certidoes.grid(row = 2, column = 0, sticky = W, padx = 27, pady = 10)
    modelosDocumentos = ttk.Button(gestaoEmpresa, text = 'Modelos de Documentos', width = 23, style = "Accentbutton", command = linkModelos)
    modelosDocumentos.grid(row = 2, column = 2, sticky = W, pady = 10)