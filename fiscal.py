from tkinter import *
from tkinter.ttk import *
import tkinter as tk
import tkinter.ttk as ttk
from PIL import ImageTk, Image
import webbrowser
from openpyxl import load_workbook

#carrega a planilha
caminho = './assets/spreadsheets/LEIAUTE PADRAO.xlsx'
caminho_icon = "./assets/images/icon-removebg-preview.ico"
caminho_logo = "./assets/images/LOGO2.png"
caminho_tema = './assets/themes/azure/azure.tcl'

arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active

def linkWppFiscal():
    wppFiscal = planilha1.cell(row = 24, column = 3).value
    webbrowser.open(wppFiscal)
def linkAreaCliente():
    areaCliente = planilha1.cell(row = 15, column = 3).value
    webbrowser.open(areaCliente)
def linkNFStock():
    nfStock = planilha1.cell(row = 26, column = 3).value
    webbrowser.open(nfStock)
def linkSintegra():
    sintegra = planilha1.cell(row = 27, column = 3).value
    webbrowser.open(sintegra)
def linkCartaoCNPJ():
    cartaoCNPJ = planilha1.cell(row = 28, column = 3).value
    webbrowser.open(cartaoCNPJ)
def linkConsultaCFOP():
    consultaCFOP = planilha1.cell(row = 29, column = 3).value
    webbrowser.open(consultaCFOP)

def jan_depFiscal():
    #cria a janela
    depFiscal = tk.Toplevel()
    depFiscal.grab_set()
    depFiscal.title("Departamento Fiscal")
    depFiscal.geometry("400x360")
    depFiscal.iconbitmap(caminho_icon)
    depFiscal.resizable(False, False)
    
    #carrega a logo
    img = ImageTk.PhotoImage(Image.open(caminho_logo))
    panel2 = tk.Label(depFiscal, image = img)
    panel2.grid(row = 0, columnspan = 3, column = 0, sticky = N, padx = 55, pady = 30)
    panel2.image = img
    
    #criando botões
    wppFiscal = ttk.Button(depFiscal, text = 'Whatsapp Dep. Fiscal', width = 23, style = "Accentbutton", command = linkWppFiscal)
    wppFiscal.grid(row = 1, column = 0, sticky = W, padx = 27, pady = 10)
    areaCliente1 = ttk.Button(depFiscal, text = 'Área do Cliente', width = 23, style = "Accentbutton", command = linkAreaCliente)
    areaCliente1.grid(row = 1, column = 2, sticky = W, pady = 10)
    nfStock = ttk.Button(depFiscal, text = 'NF-Stock', width = 23, style = "Accentbutton", command = linkNFStock)
    nfStock.grid(row = 2, column = 0, sticky = W, pady = 10, padx = 27)
    sintegra = ttk.Button(depFiscal, text = 'Sintegra', width = 23, style = "Accentbutton", command = linkSintegra)
    sintegra.grid(row = 2, column = 2, sticky = W, pady = 10)
    cartaoCNPJ = ttk.Button(depFiscal, text = 'Cartão CNPJ', width = 23, style = "Accentbutton", command = linkCartaoCNPJ)
    cartaoCNPJ.grid(row = 3, column = 0, sticky = W, pady = 10, padx = 27)
    consultaCFOP = ttk.Button(depFiscal, text = 'Consulta CFOP', width = 23, style = "Accentbutton", command = linkConsultaCFOP)
    consultaCFOP.grid(row = 3, column = 2, sticky = W, pady = 10)