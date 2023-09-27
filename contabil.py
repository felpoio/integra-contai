from tkinter import *
from tkinter.ttk import *
import tkinter as tk
import tkinter.ttk as ttk
from PIL import ImageTk, Image
import webbrowser
from openpyxl import load_workbook

#carrega a planilha
caminho = './assets/spreadsheets/LEIAUTE PADRAO.xlsx'
caminho_icon = "./assets/images/icon-removebg-preview.ico"
caminho_logo = "./assets/images/LOGO2.png"
caminho_tema = './assets/themes/azure/azure.tcl'

arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active

def linkWppContabil():
    wppContabil = planilha1.cell(row = 14, column = 3).value
    webbrowser.open(wppContabil)
def linkAreaCliente():
    areaCliente = planilha1.cell(row = 15, column = 3).value
    webbrowser.open(areaCliente)
def demonsContabil():
    demonstContabil = planilha1.cell(row = 16, column = 3).value
    webbrowser.open(demonstContabil)
def moviFinanceiro():
    moveFinanceiro = planilha1.cell(row = 17, column = 3).value
    webbrowser.open(moveFinanceiro)
def valorIdentificar():
    valueIdentificar = planilha1.cell(row = 18, column = 3).value
    webbrowser.open(valueIdentificar)
def linkGerencia():
    gerencia = planilha1.cell(row = 19, column = 3).value
    webbrowser.open(gerencia)
def linkOutros():
    outros = planilha1.cell(row = 21, column = 3).value
    webbrowser.open(outros)

def jan_depContabil():
    #cria a janela
    depContabil = tk.Toplevel()
    depContabil.grab_set()
    depContabil.title("Departamento Contábil")
    depContabil.geometry("400x415")
    depContabil.iconbitmap(caminho_icon)
    depContabil.resizable(False, False)
    
    #carrega a logo
    img = ImageTk.PhotoImage(Image.open(caminho_logo))
    panel2 = tk.Label(depContabil, image = img)
    panel2.grid(row = 0, columnspan = 3, column = 0, sticky = N, padx = 55, pady = 30)
    
    #criando botões
    wppContabil = ttk.Button(depContabil, text = 'Whatsapp Dép. Contábil', width = 23, style = "Accentbutton", command = linkWppContabil)
    wppContabil.grid(row = 1, column = 0, sticky = W, padx = 27, pady = 10)
    areaCliente1 = ttk.Button(depContabil, text = 'Área do Cliente', width = 23, style = "Accentbutton", command = linkAreaCliente)
    areaCliente1.grid(row = 1, column = 2, sticky = W, pady = 10)
    demoContabil = ttk.Button(depContabil, text = 'Demonstrações Contábeis', width = 23, style = "Accentbutton", command = demonsContabil)
    demoContabil.grid(row = 2, column = 0, sticky = W, padx = 27, pady = 10)
    movFinanceiro = ttk.Button(depContabil, text = 'Movimento Financeiro', width = 23, style = "Accentbutton", command = moviFinanceiro)
    movFinanceiro.grid(row = 2, column = 2, sticky = W, pady = 10)
    valIdentificar = ttk.Button(depContabil, text = 'Valores para Identificar', width = 23, style = "Accentbutton", command = valorIdentificar)
    valIdentificar.grid(row = 3, column = 0, sticky = W, pady = 10, padx = 27)
    gerencia = ttk.Button(depContabil, text = 'À gerência', width = 23, style = "Accentbutton", command = linkGerencia)
    gerencia.grid(row = 3, column = 2, sticky = W, pady = 10)   
    outros = ttk.Button(depContabil, text = 'Outros', width = 23, style = "Accentbutton", command = linkOutros)
    outros.grid(row = 4, column = 0, sticky = W, pady = 10, padx = 27)