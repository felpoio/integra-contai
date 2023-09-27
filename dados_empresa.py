from tkinter import *
from tkinter.ttk import *
import tkinter as tk
import tkinter.ttk as ttk
from PIL import ImageTk, Image
from openpyxl import load_workbook

#carrega a planilha
caminho = './assets/spreadsheets/LEIAUTE PADRAO.xlsx'
caminho_icon = "./assets/images/icon-removebg-preview.ico"
caminho_logo = "./assets/images/LOGO2.png"
caminho_tema = './assets/themes/azure/azure.tcl'

arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active

#definindo os links
dadoCnpj = planilha1.cell(row = 2, column = 3).value
dadoInsc_estadual = planilha1.cell(row = 3, column = 3).value
dadoInsc_municipal = planilha1.cell(row = 4, column = 3).value
razaoSocial = planilha1.cell(row = 5, column = 3).value

strInsc_municipal = str(dadoInsc_municipal)
strRazaoSocial = str(razaoSocial)

def jan_dadoEmpresa():
    #cria a janela
    dadoEmpresa = tk.Toplevel()
    dadoEmpresa.grab_set()
    dadoEmpresa.title("Dados da Empresa")
    dadoEmpresa.geometry("700x360")
    dadoEmpresa.iconbitmap(caminho_icon)
    dadoEmpresa.resizable(False, False)
    
    #carrega a logo
    img = ImageTk.PhotoImage(Image.open(caminho_logo))
    panel2 = tk.Label(dadoEmpresa, image = img)
    panel2.grid(row = 0, columnspan = 3, column = 0, sticky = N, padx = 200, pady = 30)
    
    #Conteudo da janela
    cnpj = Label(dadoEmpresa, text = 'CNPJ: ' + dadoCnpj, font=("Helvetica", 12))
    cnpj.grid(row = 1, column = 0, sticky = W, padx = 40, pady = 10)
    insc_est = Label(dadoEmpresa, text = 'Inscrição Estadual: ' + dadoInsc_estadual, font=("Helvetica", 12))
    insc_est.grid(row = 2, column = 0, sticky = W, padx = 40, pady = 10)
    insc_mun = Label(dadoEmpresa, text = 'Inscrição Municipal: ' + strInsc_municipal, font=("Helvetica", 12))
    insc_mun.grid(row = 3, column = 0, sticky = W, padx = 40, pady = 10)
    razao_social = Label(dadoEmpresa, text = 'Razão Social: ' + strRazaoSocial, font=("Helvetica", 12))
    razao_social.grid(row = 4, column = 0, sticky = W, padx = 40, pady = 10)