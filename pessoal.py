from tkinter import *
from tkinter.ttk import *
import tkinter as tk
import tkinter.ttk as ttk
from PIL import ImageTk, Image
import webbrowser
from openpyxl import load_workbook

import calculadora

#carrega a planilha
caminho = './assets/spreadsheets/LEIAUTE PADRAO.xlsx'
caminho_icon = "./assets/images/icon-removebg-preview.ico"
caminho_logo = "./assets/images/LOGO2.png"
caminho_tema = './assets/themes/azure/azure.tcl'

arquivo_excel = load_workbook(caminho)
planilha1 = arquivo_excel.active

def linkWppPessoal():
    wppPessoal = planilha1.cell(row = 32, column = 3).value
    webbrowser.open(wppPessoal)
def linkAreaCliente():
    areaCliente = planilha1.cell(row = 15, column = 3).value
    webbrowser.open(areaCliente)

def jan_depPessoal():
    #cria a janela
    depPessoal = tk.Toplevel()
    depPessoal.grab_set()
    depPessoal.title("Departamento Pessoal")
    depPessoal.geometry("400x360")
    depPessoal.iconbitmap(caminho_icon)
    depPessoal.resizable(False, False)
    
    #carrega a logo
    img = ImageTk.PhotoImage(Image.open(caminho_logo))
    panel2 = tk.Label(depPessoal, image = img)
    panel2.grid(row = 0, columnspan = 3, column = 0, sticky = N, padx = 55, pady = 30)
    
    #criando botões
    wppPessoal = ttk.Button(depPessoal, text = 'Whatsapp Dep. Pessoal', width = 23, style = "Accentbutton", command = linkWppPessoal)
    wppPessoal.grid(row = 1, column = 0, sticky = W, padx = 27, pady = 10)
    areaCliente1 = ttk.Button(depPessoal, text = 'Área do Cliente', width = 23, style = "Accentbutton", command = linkAreaCliente)
    areaCliente1.grid(row = 1, column = 2, sticky = W, pady = 10)
    calculadoraBtt = ttk.Button(depPessoal, text = 'Cálculo Funcionário', width = 23, style = "Accentbutton", command = calculadora.jan_Calculadora)
    calculadoraBtt.grid(row = 2, column = 0, sticky = W, padx = 27, pady = 10)